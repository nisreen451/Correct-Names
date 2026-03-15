from openpyxl import load_workbook
from datetime import datetime
import os

# =========================
# FILE PATH
# =========================
input_file = r"C:\Users\CFWP\Desktop\Attendance Test\UpdatedCodes\InputFiles\DetailedPayment_for_beneficaries_52202682436.xlsx"

# =========================
# NAME CORRECTION DICTIONARY
# =========================
name_corrections = {
    "IBRM": "IBRAHIM",
    "MAHM": "MAHMOUD",
    "Mahd": "MAHMOUD",
    "MAHD": "MAHMOUD",
    "A": "ABD AL",
    "CHARIF": "SHARIF",
    "CHIHAB": "SHIHAB",
    "CHLUN": "SHLUN",
    "MOHD": "MOHAMMAD",
    "Mohd": "MOHAMMAD",
    "MOHM": "MOHAMMAD",
    "MHD": "MOHAMMAD",
    "U": "ABD",
    "ABDULA": "ABDULLAH",
    "MUST": "MUSTAFA",
    "Akluk": "ALAkluk",
    "Aklouk": "ALAkluk",
    "A.": "ABD AL",
    "ARAHMAN": "ALRAHMAN",
    "HLA": "HALA",
    "WIRUD": "WUROOD",
    "ID": "EID"
}

# =========================
# CLEANING FUNCTION
# =========================
def clean_name(name):
    if name is None:
        return name
    tokens = str(name).split()
    cleaned_tokens = []
    i = 0
    while i < len(tokens):
        token = tokens[i]
        # Special merge case for "A" or "A."
        if token in ("A", "A.") and i + 1 < len(tokens):
            next_token = tokens[i + 1]
            cleaned_tokens.append("ABD AL" + next_token)
            i += 2
            continue
        # Normal replacement
        cleaned_tokens.append(name_corrections.get(token, token))
        i += 1
    return " ".join(cleaned_tokens)

# =========================
# LOAD WORKBOOK
# =========================
wb = load_workbook(input_file)

# =========================
# SELECT ONLY 'Payments' SHEET
# =========================
if 'Payments' not in wb.sheetnames:
    raise ValueError("Sheet 'Payments' not found in workbook!")

ws = wb['Payments']

# =========================
# FIND HEADER ROW AND COLUMN
# =========================
header_row = None
col_index = None
for r in range(1, 21):  # scan first 20 rows to find header
    for c in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=r, column=c).value
        if cell_value and str(cell_value).replace("\xa0"," ").strip() == "Beneficiary Name English":
            header_row = r
            col_index = c
            break
    if col_index:
        break

if not col_index:
    raise ValueError("Column 'Beneficiary Name English' not found in 'Payments' sheet!")

# =========================
# UPDATE COLUMN VALUES
# =========================
for row in range(header_row + 1, ws.max_row + 1):
    cell = ws.cell(row=row, column=col_index)
    cell.value = clean_name(cell.value)
    
# =========================
# SAVE FILE IN 'OutputFiles' FOLDER
# =========================
output_folder = r"C:\Users\CFWP\Desktop\Attendance Test\UpdatedCodes\OutputFiles"

# Ensure folder exists
if not os.path.exists(output_folder):
    raise FileNotFoundError(f"Output folder does not exist: {output_folder}")

ext = os.path.splitext(input_file)[1]  # keep same extension (.xlsx)

output_file = os.path.join(output_folder, f"payment summary with corrected names{ext}")

wb.save(output_file)
print(f"File saved successfully in:\n{output_file}")
